"""
Categories Import Endpoint - Excel (.xlsx) file upload for bulk category management
"""
from __future__ import annotations

import io
from typing import List, Dict
from fastapi import APIRouter, File, UploadFile, HTTPException, Depends
from pydantic import BaseModel
from openpyxl import load_workbook

from classymail.services.settings_store import load_settings, save_settings
from classymail.services.azure_clients import Clients, get_clients

router = APIRouter(prefix="/api/settings/categories", tags=["categories"])


class ImportResult(BaseModel):
    """Result of category import operation"""
    total_rows: int
    created: int
    updated: int
    skipped: int
    errors: List[str]
    categories: List[Dict[str, str]]


def slugify_category_name(name: str) -> str:
    """Generate URL-safe slug from category name"""
    # Remove accents and special characters
    slug = name.lower().strip()
    replacements = {
        'é': 'e', 'è': 'e', 'ê': 'e', 'ë': 'e',
        'à': 'a', 'â': 'a', 'ä': 'a',
        'î': 'i', 'ï': 'i',
        'ô': 'o', 'ö': 'o',
        'ù': 'u', 'û': 'u', 'ü': 'u',
        'ç': 'c',
        ' ': '_', '-': '_', "'": '', '"': ''
    }
    for old, new in replacements.items():
        slug = slug.replace(old, new)

    # Keep only alphanumeric and underscores
    slug = ''.join(ch for ch in slug if ch.isalnum() or ch == '_')

    # Remove duplicate underscores
    while '__' in slug:
        slug = slug.replace('__', '_')

    return slug.strip('_')


def parse_excel_categories(file_content: bytes) -> List[Dict[str, str]]:
    """Parse Excel file and extract categories

    Expected columns:
    - A: Nom de l'Intention (Category Name)
    - B: Définition (Definition)
    - C: Exclusion (Exclusions)
    """
    try:
        # Load workbook from bytes
        wb = load_workbook(filename=io.BytesIO(file_content), read_only=True)
        ws = wb.active

        # Get all rows as list
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            raise ValueError("Excel file is empty")

        # Try to find column indices by header names
        header_row = rows[0] if rows else []
        name_col_idx = None
        definition_col_idx = None
        exclusion_col_idx = None

        # Case-insensitive column matching
        for idx, cell_value in enumerate(header_row):
            if not cell_value:
                continue
            col_lower = str(cell_value).lower().strip()
            if 'nom' in col_lower and 'intention' in col_lower:
                name_col_idx = idx
            elif 'définition' in col_lower or 'definition' in col_lower:
                definition_col_idx = idx
            elif 'exclusion' in col_lower:
                exclusion_col_idx = idx

        # Fallback: use positional columns if named columns not found
        if name_col_idx is None and len(header_row) > 0:
            name_col_idx = 0
        if definition_col_idx is None and len(header_row) > 1:
            definition_col_idx = 1
        if exclusion_col_idx is None and len(header_row) > 2:
            exclusion_col_idx = 2

        if name_col_idx is None:
            raise ValueError("Cannot find 'Nom de l'Intention' column")

        categories = []

        # Skip header row and process data rows
        for row in rows[1:]:
            if not row or len(row) <= name_col_idx:
                continue

            name = str(row[name_col_idx]).strip() if row[name_col_idx] else ''

            # Skip empty rows or header-like rows
            if not name or name.lower() in ['nom', "nom de l'intention", 'name']:
                continue

            definition = ''
            if definition_col_idx is not None and len(row) > definition_col_idx and row[definition_col_idx]:
                definition = str(row[definition_col_idx]).strip()

            exclusion = ''
            if exclusion_col_idx is not None and len(row) > exclusion_col_idx and row[exclusion_col_idx]:
                exclusion = str(row[exclusion_col_idx]).strip()

            # Clean up None values
            if definition == 'None':
                definition = ''
            if exclusion == 'None':
                exclusion = ''

            categories.append({
                'name': name,
                'definition': definition,
                'exclusion': exclusion
            })

        wb.close()
        return categories

    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Error parsing Excel file: {str(e)}"
        )


@router.post("/import", response_model=ImportResult)
async def import_categories_from_excel(
    file: UploadFile = File(...),
    replace_mode: bool = False,  # If True, replace all categories; if False, merge/update
    clients: Clients = Depends(get_clients)
) -> ImportResult:
    """
    Import categories from Excel (.xlsx) file

    Expected Excel format:
    - Column A: Nom de l'Intention (Category Name)
    - Column B: Définition (Definition - what it IS)
    - Column C: Exclusion (Exclusions - what it ISN'T)

    Args:
        file: Excel file upload (.xlsx)
        replace_mode: If True, replaces all existing categories. If False, merges/updates.
        clients: Azure clients dependency

    Returns:
        ImportResult with counts and any errors
    """
    # Validate file type
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Only Excel files (.xlsx, .xls) are supported"
        )

    # Read file content
    try:
        content = await file.read()
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Error reading file: {str(e)}"
        )

    # Parse Excel
    parsed_categories = parse_excel_categories(content)

    if not parsed_categories:
        raise HTTPException(
            status_code=400,
            detail="No valid categories found in Excel file"
        )

    # Load current settings
    settings = load_settings()
    existing_categories = settings.get('categories', [])

    # Create slug → category mapping for existing categories
    existing_by_slug = {cat['slug']: cat for cat in existing_categories}
    existing_by_name = {cat['name'].lower(): cat for cat in existing_categories}

    result = ImportResult(
        total_rows=len(parsed_categories),
        created=0,
        updated=0,
        skipped=0,
        errors=[],
        categories=[]
    )

    new_categories = []

    for idx, parsed_cat in enumerate(parsed_categories, start=1):
        try:
            name = parsed_cat['name']
            definition = parsed_cat['definition']
            exclusion = parsed_cat['exclusion']

            # Generate slug
            slug = slugify_category_name(name)

            if not slug:
                result.errors.append(f"Row {idx}: Cannot generate slug for '{name}'")
                result.skipped += 1
                continue

            # Check if category exists (by slug or name)
            existing = existing_by_slug.get(slug) or existing_by_name.get(name.lower())

            category = {
                'name': name,
                'slug': slug,
                'description': definition,
                'exclusions': exclusion
            }

            if existing:
                # Update existing category
                result.updated += 1
                result.categories.append({
                    'action': 'updated',
                    'name': name,
                    'slug': slug
                })
            else:
                # Create new category
                result.created += 1
                result.categories.append({
                    'action': 'created',
                    'name': name,
                    'slug': slug
                })

            new_categories.append(category)

        except Exception as e:
            result.errors.append(f"Row {idx}: {str(e)}")
            result.skipped += 1

    # Apply changes based on replace_mode
    if replace_mode:
        # Replace all categories
        settings['categories'] = new_categories
    else:
        # Merge: keep existing categories not in import, update matching ones
        merged_slugs = {cat['slug'] for cat in new_categories}

        # Keep existing categories not in import
        kept_categories = [
            cat for cat in existing_categories
            if cat['slug'] not in merged_slugs
        ]

        # Combine: kept + new/updated
        settings['categories'] = kept_categories + new_categories

    # Save settings
    save_settings(settings)

    try:
        # Also save to Cosmos DB if available
        from classymail.services.settings_store import save_settings_async
        await save_settings_async(settings, clients=clients)
    except Exception as e:
        # Non-critical error, continue
        result.errors.append(f"Warning: Could not sync to Cosmos DB: {str(e)}")

    return result
